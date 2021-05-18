from openpyxl import load_workbook
from cts import cambioletra, eleccio_centre
from timeit import timeit
import shutil

# import socket programming library
import socket

# import thread module
from _thread import *
import threading

def escribir_excel(data):
    # Direcció de l'arxiu
    filePath = "CONTROL_DINCIDENCIES_2.xlsx"
    dades= data.split('//')
    centre = dades[0]
    causa = dades[2]
    classificacio = dades[3]
    incidencia = dades[4]
    entrada = dades[5]
    if data[1] == None:
        direccio = None
    else:
        direccio = data[1]

    wb = load_workbook(filePath)
    sheet = wb[centre]
    i: int = 0
    found0 = 0
    found1 = 0

    for row in sheet.iter_rows(min_col= 2, max_col= 2):
        for cell in row:
            if found0 == 1:
                break
            if cell.value == "CENTRE":
                found0 = 1
                minRow = cell.row
                break

    for row in sheet.iter_rows(min_row= minRow, max_row= 100, min_col= 2, max_col= 2):
        for cell in row:
            if found1 == 1:
                break
            if cell.value == None:
                found1 = 1
                RowIn = cell.row
                break

    for row in sheet.iter_rows(min_col= 1, max_col= 20):
        for cell in row:
            if cell.value == "CENTRE":
                CENTCOL = cell.column

            elif cell.value == "CLASSIFICACIÓ":
                CLASCOL = cell.column

            elif cell.value == "CAUSA":
                CAUSCOL = cell.column

            elif cell.value == "ESTIMACIÓ DIES":
                ESTCOL = cell.column

            elif cell.value == "INCIDÈNCIA/REPARACIÓ DEMANDADA":
                INCCOL = cell.column

            elif cell.value == "ENTRADA":
                ENCOL = cell.column

            elif cell.value == "DIES":
                DICOL = cell.column
                SORCOL = cell.column -1

            elif cell.value == "EF.":
                EFCOL = cell.column
                DIRCOL = cell.column + 2


    sheet[cambioletra(CENTCOL) + str(RowIn)] = eleccio_centre(centre)
    sheet[cambioletra(CLASCOL) + str(RowIn)] = classificacio
    sheet[cambioletra(CAUSCOL) + str(RowIn)] = causa
    sheet[cambioletra(INCCOL) + str(RowIn)] = incidencia
    sheet[cambioletra(ENCOL) + str(RowIn)] = entrada
    sheet[cambioletra(DIRCOL) + str(RowIn)] = direccio

    sheet[cambioletra(SORCOL) + str(RowIn)] = '=' + cambioletra(ENCOL) + str(RowIn) + "+" + cambioletra(ESTCOL) + str(RowIn)
    sheet[cambioletra(EFCOL) + str(RowIn)] = '=' + cambioletra(DICOL) + str(RowIn) + "-" + cambioletra(ESTCOL) + str(RowIn)

    wb.save(filePath)

print_lock = threading.Lock()

# thread function
def threaded(c):
    while True:
    
        #data received from client
        data = c.recv(1024)
        data = data.decode('utf-8')
        
        if not data:
             # lock released on exit
             print_lock.release()
             break


        escribir_excel(data)

        
        # send back reversed string to client --> Escriure al excel i fer un retorn diguent OK
        
    # connection closed
    c.close()
    

def Main():
    host = "10.38.133.220" # Hi ha que ficar la nostra direcció
    
    # reverse a port on your computer
    # in our case it is 12345 but it
    # can be anything
    port = 12345
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind((host , port))
    print("socket binded to port", port)
    
    # put the socket into listening mode
    s.listen(5)
    print("socket is listening")
    
    # a forever loop until client wants to exit
    while True:
        
        # establish connection with client
        c, addr = s.accept()
        #temporitzador = time()
        #if temporitzador%60 == 0:
            #shutil.copy("CONTROL_DINCIDENCIES_2.xlsx", "BackUp/CONTROL_DINCIDENCIES_2.xlsx")
        # loop acquired by client
        print_lock.acquire()
        print('Connected to :', addr[0], ':', addr[1])
        
        # Start a new thread and return its identifier
        start_new_thread(threaded, (c,))
    s.close()
    

Main()
